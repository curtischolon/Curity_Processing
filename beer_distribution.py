import gspread
from oauth2client.service_account import ServiceAccountCredentials
import curity_report_download
import os
from openpyxl import load_workbook
import time
import logging
import csv
import sys


logging.basicConfig(filename='curity.log',
                      level=logging.INFO,
                      format="%(asctime)s - %(levelname)s: %(message)s",
                      datefmt="%m/%d/%Y %I:%M:%S %p"
                          )

# clean up any lingering files from the previous run
file_list = os.listdir()
for file in file_list:
    if file.lower().endswith('error.csv'):
        os.remove(file)


csv_summary = open('curity_summary.csv', 'w', newline='')
csv_writer = csv.writer(csv_summary)
csv_writer.writerow(['Brewery', 'StartingRow', 'RowsToAppend', 'Range', 'Status', 'Sheet', 'DocID'])

logging.info('downloading curity report')
if os.path.isfile('curity_report.xlsx'):
    os.remove('curity_report.xlsx')

# begin by downloading the current curity report for processing
# the current report will have a time delta of zero days
curity_report_download.download_curity_report()

# confirm the report downloaded as expected
if os.path.isfile('curity_report.xlsx'):
    print('File downloaded')
    logging.info('curity report sucessfully downloaded')
else:
    sys.exit()

# read the curity report and log the details to the appropriate google doc
logging.info('loading curity report')
wb = load_workbook('curity_report.xlsx', read_only=True)
ws = wb.active

logging.info('extracting detail data for processing')
row_data = []
for row in range(3, ws.max_row):
    col_data = []
    logging.info(ws[row][2].value)
    for col in range(0, ws.max_column):
        col_data.append(ws[row][col].value)
    row_data.append(col_data)
wb.close()

# compile list of breweries to process
logging.info('extracting list of breweries to process')
breweries = []
for row in row_data:
    if row[2] not in breweries:
        print(row[2])
        logging.info(row[2])

        breweries.append(row[2])

# open Google API session
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
client = gspread.authorize(creds)

# Control list
# Test
# wb = load_workbook("GoogleDocControlList.xlsx")
# prod
logging.info('loading google sheets control list')
wb = load_workbook(os.path.join('\\\\bur-dvr01', 'f', 'Shared', 'RPA', 'GoogleDocs', 'GoogleDocControlList.xlsx'),
                   read_only=True)
ws = wb.active
control_list = []
row_count = 0
for row in ws.rows:
    row_count += 1
    if row_count < 2:
        continue
    control_list.append([row[0].value, row[1].value, row[3].value])
    print(row[0].value, row[1].value, row[3].value)
wb.close()

for brewery in breweries:
    # gather data for the current brewery
    cells = []
    brewery_rows = []
    rows_to_append = 0
    status = ''
    for row in range(len(row_data)):
        if row_data[row][2] == brewery:
            rows_to_append += 1
            brewery_cols = []
            for col in range(len(row_data[row])):
                cells.append(row_data[row][col])
                brewery_cols.append(row_data[row][col])
            brewery_rows.append(brewery_cols)
    logging.info(f'rows to append: {rows_to_append}')

    # get the spreadsheet name from the control list
    gs_name = ''
    gs_key = ''
    process_brewery = False
    for i in control_list:
        if i[0] and i[0].lower() == brewery.lower():
            process_brewery = True
            gs_name = i[1]
            gs_key = i[2]

    if not process_brewery:
        logging.info(f'{brewery} is not in the control sheet')
        status = 'Brewery not in control doc'
        csv_writer.writerow([brewery,
                             '',
                             rows_to_append,
                             '',
                             status,
                             '',
                             ''])
        with open(f'{brewery}_error.csv', 'w', newline='') as error_out:
            csv_error_writer = csv.writer(error_out)
            print(brewery_rows)
            for row in brewery_rows:
                print(row)
                csv_error_writer.writerow(row)
        continue

    # find workbook
    logging.info(f'processing: {brewery}')
    print(brewery)
    # sheet = client.open(gs_name)
    sheet = client.open_by_key("1YWDcfIJ_Q9n-iHsd1L9bWhzVuTp1V13kU3XSKlrG0PI")
    logging.info('connecting to google sheet')
    sheet = client.open_by_key(gs_key)
    try:
        ws = sheet.get_worksheet(0)
        logging.info('sucessfully connected to google sheet')
    except gspread.exceptions.APIError as e:
        print("No available file for the current brewery: {}".format(brewery))
        status = 'Unable to connect to sheet'
        logging.info('google doc is not available for the current brewery')
        csv_writer.writerow([brewery,
                             '',
                             rows_to_append,
                             '',
                             status,
                             '',
                             gs_key])
        with open(f'{brewery}_error.csv', 'w', newline='') as error_out:
            csv_error_writer = csv.writer(error_out)
            print(brewery_rows)
            for row in brewery_rows:
                print(row)
                csv_error_writer.writerow(row)
        continue

    try:
        all_cells = ws.get_all_values()
    except gspread.exceptions.APIError as e:
        print(e)
        logging.warning(f'API exception logged for {brewery}')
        time.sleep(100)
        all_cells = ws.get_all_values()
    starting_row = len(all_cells) + 1
    logging.info(f'starting row: {starting_row}')

    ws.add_rows(rows_to_append)
    ws_range = ws.range('A{}:R{}'.format(starting_row, starting_row + rows_to_append))
    logging.info(f'range: A{starting_row}:R{starting_row+rows_to_append}')
    for cell in range(len(cells)):
        ws_range[cell].value = cells[cell]

    # write data to the gs
    ws.update_cells(ws_range, value_input_option='USER_ENTERED')
    csv_writer.writerow([brewery,
                         starting_row,
                         rows_to_append,
                         f'A{starting_row}:R{starting_row + rows_to_append}',
                         'Processed successfully',
                         ws.title,
                         gs_key]
                        )
    time.sleep(2)
csv_summary.close()